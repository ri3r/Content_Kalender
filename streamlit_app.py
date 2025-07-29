import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random
import io
import requests
import time

def generate_content_openai(prompt, api_key, model, max_tokens=80, temperature=0.7, retries=3, retry_delay=1):
    prompt = prompt.strip()
    if not prompt.endswith("Bitte antworte auf Deutsch."):
        prompt += " Bitte antworte auf Deutsch."
    endpoint = "https://api.openai.com/v1/chat/completions"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": temperature
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    for attempt in range(retries):
        try:
            response = requests.post(endpoint, headers=headers, json=payload, timeout=15)
            response.raise_for_status()
            content = response.json()["choices"][0]["message"]["content"].strip()
            content = content.strip().strip('"').strip("'")
            return content
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(retry_delay)
            else:
                st.error(f"Fehler bei der OpenAI-Anfrage: {e}")
                return "Idee konnte nicht automatisch generiert werden."

def generate_date_range(start_date, num_days):
    return [start_date + timedelta(days=i) for i in range(num_days)]

def create_excel_calendar(df, customer, content_formats, status_options, topic_options):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content Kalender"

    # Schreibe die Header
    for col, header in enumerate(df.columns, 1):
        ws.cell(row=1, column=col).value = header

    # Schreibe die Daten
    for row, record in enumerate(df.itertuples(index=False), 2):
        for col, value in enumerate(record, 1):
            ws.cell(row=row, column=col).value = value

    # -- Schreibe Drop-Down-Werte in ein separates Blatt --
    dv_sheet = wb.create_sheet(title="Dropdowns")
    # Themen
    for i, val in enumerate(topic_options, 1):
        dv_sheet.cell(row=i, column=1).value = val
    # Content-Format
    for i, val in enumerate(content_formats, 1):
        dv_sheet.cell(row=i, column=2).value = val
    # Status
    for i, val in enumerate(status_options, 1):
        dv_sheet.cell(row=i, column=3).value = val

    # Definiere Bereichs-Formeln
    topic_range = f"Dropdowns!$A$1:$A${len(topic_options)}"
    content_format_range = f"Dropdowns!$B$1:$B${len(content_formats)}"
    status_range = f"Dropdowns!$C$1:$C${len(status_options)}"

    # Drop-Down für Content-Format
    if "Content-Format" in df.columns:
        colidx = df.columns.get_loc("Content-Format") + 1
        dv = DataValidation(type="list", formula1=f"={content_format_range}", allow_blank=True)
        dv.add(f"{openpyxl.utils.get_column_letter(colidx)}2:{openpyxl.utils.get_column_letter(colidx)}{len(df)+1}")
        ws.add_data_validation(dv)
    # Drop-Down für Status
    if "Status" in df.columns:
        colidx = df.columns.get_loc("Status") + 1
        dv = DataValidation(type="list", formula1=f"={status_range}", allow_blank=True)
        dv.add(f"{openpyxl.utils.get_column_letter(colidx)}2:{openpyxl.utils.get_column_letter(colidx)}{len(df)+1}")
        ws.add_data_validation(dv)
    # Drop-Down für Thema
    if "Thema" in df.columns:
        colidx = df.columns.get_loc("Thema") + 1
        dv = DataValidation(type="list", formula1=f"={topic_range}", allow_blank=True)
        dv.add(f"{openpyxl.utils.get_column_letter(colidx)}2:{openpyxl.utils.get_column_letter(colidx)}{len(df)+1}")
        ws.add_data_validation(dv)

    # Dropdown-Blatt ausblenden
    dv_sheet.sheet_state = 'hidden'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.title("Content Kalender Generator")

with st.expander("Was macht dieses Tool?"):
    st.markdown("""
    - **Alle Posts für alle Plattformen in EINEM Kalender untereinander, Plattformen gemischt**
    - Strategie-Themen, Status & Content-Format als Drop-Downs in Excel (Profi-Variante)
    - Alles flexibel anpassbar (Plattformen, Themen, Content-Formate, Status etc.)
    - Automatische Content-Ideen per OpenAI (GPT) für Social Media
    """)

with st.expander("Wie prompte ich in dem Tool?"):
    st.markdown("""
    **Tipps für Prompts:**
    - Nutze die Platzhalter **{platform}** (z.B. Instagram), **{post_type}** (z.B. Feed-Post mit Visual), **{theme}** (z.B. Strategie-Thema).
    - Je konkreter, desto besser: Thema, Zielgruppe, Ton, Länge.
    - Beispiel:  
      *„Schreibe eine kreative Social Media-Idee für {platform} als {post_type} zum Thema {theme}. Max. 50 Wörter.“*
    """)

api_key = st.text_input(
    "OpenAI API Key", type="password",
    help="API Key von https://platform.openai.com/api-keys"
)
model = st.selectbox(
    "OpenAI Modell", ["gpt-3.5-turbo", "gpt-4o", "gpt-4"],
    help="GPT-4o liefert oft kreativere Ideen."
)
customer = st.text_input("Kundenname", "Gino's Espressohaus")
num_days = st.number_input(
    "Zeitraum (Tage)", min_value=30, max_value=365, value=90,
    help="Wie viele Tage soll der Kalender umfassen?"
)
start_date = st.date_input("Startdatum", value=datetime.today())

# ----- Plattformen -----
st.markdown("### Plattformen")
if "platforms" not in st.session_state:
    st.session_state.platforms = ["Instagram", "Facebook", "TikTok", "LinkedIn"]

def add_platform():
    plat = st.session_state["new_platform_input"].strip()
    if not plat:
        st.warning("Der Name der Plattform darf nicht leer sein.")
    elif plat in st.session_state.platforms:
        st.warning("Diese Plattform gibt es schon!")
    else:
        st.session_state.platforms.append(plat)
    st.session_state["new_platform_input"] = ""

add_col, btn_col = st.columns([4, 1])
with add_col:
    st.text_input(
        "Plattform hinzufügen",
        value="",
        key="new_platform_input",
        placeholder="z.B. Pinterest"
    )
with btn_col:
    st.button("➕", key="add_platform_btn", on_click=add_platform)

if st.session_state.platforms:
    for i, plat in enumerate(st.session_state.platforms):
        col1, col2 = st.columns([8, 1])
        col1.write(plat)
        if col2.button("❌", key=f"delete_platform_{i}"):
            st.session_state.platforms.pop(i)
            st.rerun()
else:
    st.info("Noch keine Plattformen eingetragen.")

st.markdown("---")

# ----- Content-Formate -----
st.markdown("### Content-Formate")
if "content_formats" not in st.session_state:
    st.session_state.content_formats = [
        "Feed-Post mit Visual", "Feed-Post Karussel", "Feed-Post mit Video", "Story"
    ]

def add_content_format():
    fmt = st.session_state["new_content_format_input"].strip()
    if fmt and fmt not in st.session_state.content_formats:
        st.session_state.content_formats.append(fmt)
    st.session_state["new_content_format_input"] = ""

cf_add_col, cf_btn_col = st.columns([4, 1])
with cf_add_col:
    st.text_input("Content-Format hinzufügen", key="new_content_format_input", placeholder="z.B. Umfrage")
with cf_btn_col:
    st.button("➕", key="add_content_format_btn", on_click=add_content_format)

for i, fmt in enumerate(st.session_state.content_formats):
    col1, col2 = st.columns([8,1])
    col1.write(fmt)
    if col2.button("❌", key=f"del_content_format_{i}"):
        st.session_state.content_formats.pop(i)
        st.rerun()

st.markdown("---")

# ----- Status-Optionen -----
st.markdown("### Status-Optionen")
if "status_options" not in st.session_state:
    st.session_state.status_options = [
        "in Planung", "erledigt", "in Bearbeitung", "eingeplant", "on hold", "freigegeben"
    ]

def add_status_option():
    opt = st.session_state["new_status_option_input"].strip()
    if opt and opt not in st.session_state.status_options:
        st.session_state.status_options.append(opt)
    st.session_state["new_status_option_input"] = ""

so_add_col, so_btn_col = st.columns([4, 1])
with so_add_col:
    st.text_input("Status hinzufügen", key="new_status_option_input", placeholder="z.B. Review")
with so_btn_col:
    st.button("➕", key="add_status_option_btn", on_click=add_status_option)

for i, stat in enumerate(st.session_state.status_options):
    col1, col2 = st.columns([8,1])
    col1.write(stat)
    if col2.button("❌", key=f"del_status_option_{i}"):
        st.session_state.status_options.pop(i)
        st.rerun()

st.markdown("---")

# ----- Themen/Topics -----
st.markdown("### Strategie-Themen")
default_topics = [
    "Operative Effizienz und Prozessoptimierung",
    "Compliance, Sicherheit und Risikomanagement",
    "Benutzerfreundlichkeit und Zukunftsfähigkeit",
    "Reverse Engineering",
    "Webinare & Events",
    "Lebensmittelindustrie",
    "Chemieindustrie",
    "Pharmaindustrie",
    "Branchenübergreifend: Digitalisierung",
    "Branchenübergreifend: Nachhaltigkeit",
    "Branchenübergreifend: Lieferkette/Trends",
    "Branchenübergreifend: Vegan/Labels",
    "Kundenreferenz / Success Story"
]
if "topic_options" not in st.session_state:
    st.session_state.topic_options = default_topics.copy()

def add_topic_option():
    topic = st.session_state["new_topic_option_input"].strip()
    if topic and topic not in st.session_state.topic_options:
        st.session_state.topic_options.append(topic)
    st.session_state["new_topic_option_input"] = ""

to_add_col, to_btn_col = st.columns([4, 1])
with to_add_col:
    st.text_input("Strategie-Thema hinzufügen", key="new_topic_option_input", placeholder="z.B. Nachhaltige Verpackung")
with to_btn_col:
    st.button("➕", key="add_topic_option_btn", on_click=add_topic_option)

for i, t in enumerate(st.session_state.topic_options):
    col1, col2 = st.columns([8,1])
    col1.write(t)
    if col2.button("❌", key=f"del_topic_option_{i}"):
        st.session_state.topic_options.pop(i)
        st.rerun()

st.markdown("---")

# ----- Themen & Beispiele für die Promptlogik (wie gehabt) -----
st.markdown("### Themen-Prompts & Beispielideen")
if "themes" not in st.session_state:
    st.session_state.themes = [
        {
            "name": "Kunden Prompt",
            "prompt": "Schreibe eine kreative Social Media-Idee für {platform} als {post_type} zum Thema {theme}.",
            "examples": ["Führung Würzburg", "Wanderroute Festung"]
        }
    ]

if st.button("Beispiel-Daten laden"):
    st.session_state.themes = [
        {"name": "Würzburg", "prompt": "Schreibe eine kreative Social Media-Idee für {platform} als {post_type} zum Thema {theme}.", "examples": ["Führung Volkach", "Wanderroute Prichsenstadt"]},
        {"name": "Recruiting", "prompt": "Schreibe eine Recruiting-Content-Idee für {platform} als {post_type} zum Thema {theme}.", "examples": ["Stellenanzeige Azubi", "Karriere bei uns"]},
        {"name": "Finanztipp", "prompt": "Erstelle einen Finanz-Tipp-Post für {platform} als {post_type} zum Thema {theme}.", "examples": ["Spar-Tipp", "ETF einfach erklärt"]},
    ]

del_theme_idx = None
for i, theme in enumerate(st.session_state.themes):
    cols = st.columns([3,3,3,1])
    theme["name"] = cols[0].text_input(
        f"Themenname {i+1}", value=theme["name"], key=f"name_{i}", help="Kurzer Themen-Titel"
    )
    theme["prompt"] = cols[1].text_area(
        f"Prompt {i+1}", value=theme["prompt"], key=f"prompt_{i}", height=60,
        help=(
            "So gibst du OpenAI eine Aufgabe für das Thema vor. Nutze die Platzhalter {platform}, {post_type}, {theme}."
        )
    )
    theme["examples"] = [
        ex.strip() for ex in cols[2].text_area(
            f"Beispiel-Ideen {i+1} (kommagetrennt)", value=", ".join(theme["examples"]),
            key=f"ex_{i}", help="Mehrere Beispiel-Ideen, getrennt durch Kommas."
        ).split(",") if ex.strip()
    ]
    if cols[3].button("❌", key=f"del_theme_{i}", help="Dieses Thema löschen"):
        del_theme_idx = i
if del_theme_idx is not None:
    st.session_state.themes.pop(del_theme_idx)

new_theme_name = st.text_input(
    "Neues Prompt-Muster hinzufügen (interner Name)", value="", key="new_theme_name", help="Neuer Prompt-Titel"
)
new_theme_prompt = st.text_input(
    "Prompt für dieses Muster", value="", key="new_theme_prompt",
    help="Nutze die Platzhalter {platform}, {post_type}, {theme} für maximalen Kontext."
)
new_theme_examples = st.text_input(
    "Beispiele (kommagetrennt, optional)", value="", key="new_theme_examples",
    help="Beispiele für dieses Prompt-Muster"
)

if st.button("Prompt-Muster hinzufügen"):
    name = new_theme_name.strip()
    if not name:
        st.warning("Das Prompt-Muster braucht einen Namen.")
    elif any(t["name"] == name for t in st.session_state.themes):
        st.warning("Prompt-Muster existiert schon!")
    else:
        ex_list = [ex.strip() for ex in new_theme_examples.split(",") if ex.strip()]
        st.session_state.themes.append({
            "name": name,
            "prompt": new_theme_prompt.strip() if new_theme_prompt.strip() else f"Schreibe eine kreative Social Media-Idee für {{platform}} als {{post_type}} zum Thema {{theme}}.",
            "examples": ex_list or [name]
        })
        st.session_state["new_theme_name"] = ""
        st.session_state["new_theme_prompt"] = ""
        st.session_state["new_theme_examples"] = ""
        st.rerun()

st.markdown("---")

# ----- Frequenzen pro Plattform -----
st.markdown("### Wöchentliche Frequenz je Plattform")
# Einzigartige Plattformen, keine leeren, keine Duplikate
platforms_clean = []
for p in st.session_state.platforms:
    if p and p not in platforms_clean:
        platforms_clean.append(p)

frequencies = {}
if platforms_clean:
    cols = st.columns(len(platforms_clean))
    for idx, p in enumerate(platforms_clean):
        frequencies[p] = cols[idx].number_input(
            f"{p}",
            min_value=0,
            max_value=7,
            value=2,
            key=f"freq_{p}_{idx}",
            help=f"Wie oft pro Woche soll auf {p} gepostet werden?"
        )
else:
    st.info("Noch keine Plattformen eingetragen.")

st.markdown("---")

# --------- Kalender generieren (Plattformen gemischt!) ---------
if st.button("Kalender generieren"):
    if not api_key.strip():
        st.warning("Bitte gib einen gültigen OpenAI API Key ein!")
    elif not platforms_clean:
        st.warning("Mindestens eine Plattform ist erforderlich.")
    elif not st.session_state.themes or not all(t["name"] for t in st.session_state.themes):
        st.warning("Mindestens ein Prompt-Muster ist erforderlich.")
    elif not st.session_state.content_formats:
        st.warning("Mindestens ein Content-Format ist erforderlich.")
    elif not st.session_state.status_options:
        st.warning("Mindestens eine Status-Option ist erforderlich.")
    elif not st.session_state.topic_options:
        st.warning("Mindestens ein Strategie-Thema ist erforderlich.")
    else:
        st.info("Bitte warte. Die Content-Ideen werden jetzt per OpenAI generiert ...")
        date_list = generate_date_range(start_date, num_days)
        rows = []
        topic_count = len(st.session_state.topic_options)
        topic_idx = 0

        posting_days = [d for d in date_list if d.weekday() < 5]
        plan = []
        for p in platforms_clean:
            freq = frequencies[p]
            if freq == 0:
                continue
            n_posts = freq * (num_days // 7)
            sel_days = posting_days[::max(1, len(posting_days)//n_posts)][:n_posts]
            for d in sel_days:
                plan.append({"date": d, "platform": p})

        plan = sorted(plan, key=lambda x: (x["date"], x["platform"]))

        total = len(plan)
        done = 0
        progress = st.progress(0) if total > 0 else None

        for idx, entry in enumerate(plan):
            date = entry["date"]
            p = entry["platform"]
            topic = st.session_state.topic_options[topic_idx % topic_count]
            topic_idx += 1
            theme_obj = random.choice(st.session_state.themes)
            content_format = random.choice(st.session_state.content_formats)
            status = st.session_state.status_options[0]
            prompt = theme_obj["prompt"] \
                .replace("{platform}", p) \
                .replace("{post_type}", content_format) \
                .replace("{theme}", topic)
            if api_key.strip():
                content = generate_content_openai(prompt, api_key, model)
            else:
                content = random.choice(theme_obj["examples"])
            rows.append([
                date.strftime("%d.%m.%Y"),
                date.isocalendar()[1],
                date.strftime("%A"),
                p,
                topic,
                content_format,
                content,
                status
            ])
            done += 1
            if progress:
                progress.progress(done / total)
        if progress:
            progress.empty()
        if not rows:
            st.warning("Keine Einträge erzeugt. Bitte prüfe Plattformen, Frequenzen und Zeitraum.")
        else:
            df = pd.DataFrame(
                rows,
                columns=[
                    "Datum", "KW", "Tag", "Plattform",
                    "Thema", "Content-Format", "Inhalt", "Status"
                ]
            )
            st.success("Fertig! Du kannst die Tabelle als Excel oder CSV exportieren.")
            st.dataframe(df)
            excel_file = create_excel_calendar(
                df, customer,
                st.session_state.content_formats,
                st.session_state.status_options,
                st.session_state.topic_options
            )
            st.download_button(
                label="Content Kalender als Excel herunterladen",
                data=excel_file,
                file_name=f"Content_Kalender_{customer}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.download_button(
                label="Content Kalender als CSV herunterladen",
                data=df.to_csv(index=False, sep=";").encode("utf-8-sig"),
                file_name=f"Content_Kalender_{customer}.csv",
                mime="text/csv"
            )
